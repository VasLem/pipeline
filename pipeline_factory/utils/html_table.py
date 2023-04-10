from urllib.request import pathname2url
import os
from bs4 import BeautifulSoup
from jinja2 import Environment, BaseLoader
import pandas as pd
from typing import Optional


HTML_TABLE_TEMPLATE = """
<table id="table_id" class="display" border="0" cellspacing="0" cellpadding="0">
  <!-- HEADER ROW -->
  <thead><tr>
    {% for col in range(1, sheet.max_column + 1): %}
    <th>{{ sheet.cell(1, col).value }}</th>
    {% endfor %}
  </tr></thead>
 
  <!-- DATA ROWS -->
  <tbody>
  {% for row in range(2, sheet.max_row + 1) %}
  <tr>
    {% for col in range(1, sheet.max_column + 1): %}
    <td>{{ sheet.cell(row, col).value }}</td>
    {% endfor %}
  </tr>
  {% endfor %}
  </tbody>
</table>
"""


def makeHtmlTable(
    report: BeautifulSoup,
    inputExcelPath: str,
    outputHTMLPath: str,
    relPath: str,
    sheetName: Optional[str] = None,
):
    """
    The makeHtmlTable function creates an HTML table from a given Excel file.
    It takes three arguments:
        report - the BeautifulSoup object containing the report markup; this is used to add the table to it.
        inputExcelPath - a string representing the path of an Excel file that contains data for creating a table.

    Args:
        report:BeautifulSoup: Create the html report
        inputExcelPath:str: Specify the path to the excel file that contains the data used in generating this report
        outputHTMLPath:str: Specify the path to which the html file will be written
        relPath: str: The output path relative to the html report
        sheetName:str=None: Specify the sheet name in the excel file

    Returns:
        An iframe containing the html table.

    """

    import openpyxl

    book = openpyxl.load_workbook(inputExcelPath)
    sheet = book.active if sheetName is None else book[sheetName]

    rtemplate = Environment(loader=BaseLoader).from_string(HTML_TABLE_TEMPLATE)  # type: ignore
    htmlTable = rtemplate.render(sheet=sheet).replace("None", "")

    htmlTable = BeautifulSoup(htmlTable, "html.parser")

    widths = []
    for k, cd in sheet.column_dimensions.items():  # type: ignore
        widths.append(cd.width)
    endRow = sheet.max_row  # type: ignore
    heights = [sheet.row_dimensions[i].height for i in range(endRow)]  # type: ignore

    for cnt, tr in enumerate(htmlTable.find_all("tr")):
        for t, td in enumerate(tr.find_all("td")):
            td.attrs[
                "style"
            ] = "border-bottom-style: solid;border-bottom-width: 1px;border-collapse: collapse;border-left-style: solid;border-left-width: 1px;border-right-style: solid;border-right-width: 1px;border-top-style: solid;border-top-width: 1px;font-size: 11.0px;font-weight: bold;height: 19pt;text-align: center"

    for th in htmlTable.find_all("th"):
        th.attrs[
            "style"
        ] = "border-bottom-style: solid;border-bottom-width: 1px;border-collapse: collapse;border-left-style: solid;border-left-width: 1px;border-right-style: solid;border-right-width: 1px;border-top-style: solid;border-top-width: 1px;font-size: 11.0px;font-weight: bold;height: 19pt;text-align: center"

    for tr in htmlTable.find_all("tr"):
        if not tr.text:
            tr.decompose()

    table = htmlTable.find("table")
    if table is None:
        raise ValueError("No table found, probably invalid excel provided.")
    colgroup = htmlTable.new_tag("colgroup")
    for w in widths:
        colgroup.append(htmlTable.new_tag("col", style=f"width: {w}px"))
    table.insert(0, colgroup)

    outputHtml = BeautifulSoup()
    html = outputHtml.new_tag("html")
    body = outputHtml.new_tag("body")

    dataTableStyle = outputHtml.new_tag(
        "link",
        rel="stylesheet",
        type="text/css",
        href="https://cdn.datatables.net/1.13.1/css/jquery.dataTables.min.css",
    )
    prescript = outputHtml.new_tag(
        "script",
        type="text/javascript",
        charset="utf8",
        src="https://cdn.datatables.net/1.13.1/js/jquery.dataTables.min.js",
    )
    postscript = outputHtml.new_tag("script")
    postscript.append(
        """
$(document).ready( function () {
    $('#table_id').DataTable();
} );
        """
    )

    body.append(htmlTable)
    jQuery = report.new_tag(
        "script",
        src="https://code.jquery.com/jquery-3.6.1.min.js",
    )

    html.append(jQuery)
    html.append(dataTableStyle)
    html.append(prescript)
    html.append(body)
    html.append(postscript)
    outputHtml.append(html)
    with open(outputHTMLPath, "w") as out:
        out.write(str(outputHtml))
    nrows = pd.read_excel(
        inputExcelPath,
        sheet_name=sheetName if sheetName else 0,
    ).shape[0]
    frame = report.new_tag(
        "iframe",
        attrs={
            "height": str(
                min(
                    nrows * 16,
                    500,
                )
            )
            + "px",
            "class": "resizable",
            "src": pathname2url(relPath),
        },
    )
    return frame
