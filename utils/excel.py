import os
import time
from typing import Union

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from utils.logging import LOGGER


def saveDfToSheet(
    df: pd.DataFrame,
    writer: Union[pd.ExcelWriter, str],
    sheetName: str = "Sheet1",
    alignCenter: bool = True,
    autoAdjustColumnWidth: bool = True,
    **kwargs,
):
    """
    Save dataframe to sheet name using given writer and sheetName.

    Args:
        writer (Union[pd.ExcelWriter, str]): The excel writer, or the path to save the dataframe
        df (pd.DataFrame): the dataframe
        sheetName (str): the sheetName. Defaults to Sheet1
        alignCenter (bool, optional): Center align all columns. Defaults to true.
        autoAdjustWidth (bool, optional): Auto adjust column width. Defaults to True.
    """

    def align_center(x):
        return ["text-align: center" for x in x]

    createWriter = isinstance(writer, str)
    if createWriter:
        try:
            while True:
                if not os.path.isfile(writer):
                    raise FileNotFoundError
                try:
                    wb = openpyxl.load_workbook(writer)
                except BaseException:
                    LOGGER.warning(f"Found unreadable file {writer}, removing..")
                    os.remove(writer)
                    raise
                try:
                    wb.active
                    if sheetName in wb.sheetnames:
                        if len(wb.sheetnames) == 1:
                            raise FileNotFoundError
                        else:
                            wb.remove(wb[sheetName])
                            wb.save(writer)
                    break
                except FileNotFoundError:
                    raise
                except BaseException:
                    LOGGER.warning(
                        f"Workbook {writer} open in Excel, waiting for you to close it.."
                    )
                    time.sleep(1)
                finally:
                    try:
                        wb.close()
                    except BaseException:
                        pass

            writer = pd.ExcelWriter(writer, engine="openpyxl", mode="a")
        except BaseException:
            while True:
                try:
                    writer = pd.ExcelWriter(writer, engine="openpyxl", mode="w")
                    break
                except PermissionError:
                    LOGGER.warning(
                        f"Workbook {writer} open in Excel, waiting for you to close it.."
                    )
                    time.sleep(1)
    if alignCenter:
        df.style.apply(align_center, axis=0).to_excel(
            writer, sheet_name=sheetName, **kwargs
        )
    else:
        df.to_excel(writer, sheet_name=sheetName, **kwargs)
    if autoAdjustColumnWidth:
        for column in df:
            column_length = max(df[column].astype(str).map(len).max(), len(str(column)))
            col_idx = df.columns.get_loc(column) + 1
            writer.sheets[sheetName].column_dimensions[
                get_column_letter(col_idx)
            ].width = column_length
    if createWriter:
        writer.close()
